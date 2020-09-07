import openpyxl
import openpyxl as xl 
from openpyxl import Workbook
from openpyxl.workbook import Workbook
from openpyxl.styles import Alignment
from openpyxl.cell import Cell
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors, Fill, fills
import random
import os
#Load  workbook
workbook = openpyxl.load_workbook('out.xlsx')
worksheet_out = workbook.worksheets[0]
worksheet_out = workbook.active
worksheet = workbook.worksheets[1]
mr = worksheet.max_row
mc = worksheet.max_column
#get time
#1 Get num_les
col_stt = 'A2'
i = 0
while (worksheet[col_stt].value is not None ):
    i= i+1
    col_stt = 'A' +  str(i+1)
    cell_null = worksheet[col_stt].value
num_les=i+1

#2 get timeline
for i in range (2,num_les) :
    coler0 = 'I'+ str(i)
    coler1 = 'K'+ str(i)
    coler2 = 'L'+ str(i) 
    coler3 = 'M'+ str(i) 
    worksheet[coler1] = '=MONTH(' + coler0 + ')'
    worksheet[coler2] = '=DAY(' + coler0 + ')'
    worksheet[coler3] = '=C'+str(i)+'&CHAR(10)&J'+str(i)
workbook.save("out.xlsx")
os.system('python convertFormulatoText.py')
#ReLoad  workbook
workbook = openpyxl.load_workbook('out.xlsx')
worksheet = workbook.worksheets[1]
#Merge table
mr0 = worksheet_out.max_row
mc0 = worksheet_out.max_column
## identify for col 
for i in range (4, 10):
    workbook.worksheets[0].cell(row = 111, column = i ).value = 'T' + str(i-2)
## Merging
data_color = ['E0699C','E0D675','C25DE0','48E04E','7253E0','69ADE0','E07875','5DE0D3','E0A448','53E082']
for row in  range (2, num_les):
    scell = worksheet.cell(row = row, column = 11 ).value +3
    ecell =  worksheet.cell(row = row, column = 12 ).value +3
    col_day = worksheet.cell(row = row, column = 8 ).value
    for checker in range (4,10):
        cell_checker = worksheet_out.cell(row = 111, column = checker).value
        if  cell_checker == col_day :
            workbook.worksheets[0].merge_cells(start_row= scell, start_column= checker, end_row= ecell , end_column= checker)
            workbook.worksheets[0].cell(row = scell, column = checker).value = workbook.worksheets[1].cell(row = row, column = 13).value
            top_left_cell = workbook.worksheets[0].cell(row = scell, column = checker)
            top_left_cell.fill = PatternFill("gray125", fgColor= random.choice(data_color))
#save
workbook.save("out.xlsx")
