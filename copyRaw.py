# Copy Raw file xlsx
from path import getpath
import path
import os
# importing openpyxl module 
import openpyxl as xl 

# opening the source excel file 
filename =path.getpath()
workbook_source = xl.load_workbook(filename) 
worksheet_source = workbook_source.worksheets[0] 


# opening the destination excel file 
filename1 ="out.xlsx"
workbook_dest = xl.load_workbook(filename1)  
worksheet_dest = workbook_dest.active 
worksheet_dest = workbook_dest.worksheets[1]
# calculate total number of rows and 
# columns in source excel file 
mr = worksheet_source.max_row 
mc = worksheet_source.max_column 

# copying the cell values from source 
# excel file to destination excel file 
for i in range (1, mr + 1): 
	for j in range (1, mc + 1): 
		# reading cell value from source excel file 
		c = worksheet_source.cell(row = i, column = j) 

		# writing the read value to destination excel file 
		worksheet_dest.cell(row = i, column = j).value = c.value 

# saving the destination excel file 
workbook_dest.save(str(filename1)) 